from flask import Flask, request, redirect, url_for, render_template_string, session, abort, send_file
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from io import BytesIO
from openpyxl import Workbook, load_workbook
import sqlite3
import threading
import time
import base64
import qrcode

app = Flask(__name__)
app.secret_key = "seat-demo-secret-key-change-me"
app.config["MAX_CONTENT_LENGTH"] = 2 * 1024 * 1024

DB_PATH = "seat_reservation.db"
DEFAULT_SEATS = [
    ("A-01", "A区1号工位"), ("A-02", "A区2号工位"), ("A-03", "A区3号工位"), ("A-04", "A区4号工位"),
    ("B-01", "B区1号工位"), ("B-02", "B区2号工位"), ("B-03", "B区3号工位"), ("B-04", "B区4号工位"),
    ("C-01", "C区1号工位"), ("C-02", "C区2号工位"), ("C-03", "C区3号工位"), ("C-04", "C区4号工位"),
]
LOCK = threading.RLock()

BASE_CSS = """
<style>
:root {
  --bg: #f5f7fb;
  --panel: #ffffff;
  --panel-soft: #fafbff;
  --text: #111827;
  --muted: #6b7280;
  --border: #e5e7eb;
  --border-strong: #d8dee8;
  --primary: #111827;
  --primary-soft: #eef2ff;
  --success: #065f46;
  --danger: #dc2626;
  --shadow-sm: 0 2px 8px rgba(15, 23, 42, 0.04);
  --shadow-md: 0 10px 30px rgba(15, 23, 42, 0.08);
  --radius-sm: 12px;
  --radius-md: 18px;
  --radius-lg: 24px;
}
* { box-sizing: border-box; }
html { scroll-behavior: smooth; }
body {
  margin: 0;
  font-family: Arial, Helvetica, sans-serif;
  background:
    radial-gradient(circle at top left, #ffffff 0%, #f5f7fb 35%),
    linear-gradient(180deg, #f8fafc 0%, #f4f7fb 100%);
  color: var(--text);
}
.container { max-width: 1240px; margin: 0 auto; padding: 36px 24px 56px; }
.auth-container { max-width: 560px; margin: 64px auto; padding: 0 20px; }
.topbar, .page-head { display: flex; justify-content: space-between; align-items: flex-start; gap: 16px; margin-bottom: 24px; flex-wrap: wrap; }
.title { font-size: 34px; font-weight: 800; margin: 10px 0 8px; letter-spacing: -0.03em; line-height: 1.15; }
.subtitle { color: var(--muted); line-height: 1.7; max-width: 920px; font-size: 15px; }
.pill {
  display: inline-flex;
  align-items: center;
  gap: 8px;
  background: rgba(255,255,255,0.75);
  backdrop-filter: blur(8px);
  border: 1px solid var(--border);
  padding: 8px 14px;
  border-radius: 999px;
  font-size: 13px;
  color: #374151;
  box-shadow: var(--shadow-sm);
}
.stats { display: grid; grid-template-columns: repeat(3, 1fr); gap: 16px; margin-bottom: 24px; }
.card {
  background: rgba(255,255,255,0.9);
  backdrop-filter: blur(6px);
  border: 1px solid rgba(229,231,235,0.9);
  border-radius: var(--radius-md);
  box-shadow: var(--shadow-sm);
  transition: transform .18s ease, box-shadow .18s ease, border-color .18s ease;
}
.card:hover { transform: translateY(-3px); box-shadow: 0 14px 36px rgba(15, 23, 42, 0.1); border-color: var(--border-strong); }
.stat-card { padding: 26px; min-height: 170px; display: flex; flex-direction: column; justify-content: center; }
.stat-label { color: var(--muted); font-size: 14px; }
.stat-value { font-size: 46px; font-weight: 800; margin-top: 10px; letter-spacing: -0.04em; line-height: 1; }
.seat-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; }
.seat-card { padding: 22px; }
.seat-head { display: flex; justify-content: space-between; align-items: center; gap: 10px; margin-bottom: 14px; }
.seat-title { font-size: 24px; font-weight: 800; letter-spacing: -0.03em; line-height: 1.15; }
.badge { font-size: 12px; padding: 7px 11px; border-radius: 999px; border: 1px solid var(--border-strong); background: #f9fafb; }
.badge.busy { background: #fff7ed; border-color: #fdba74; color: #9a3412; }
.badge.free { background: #f3f4f6; color: #374151; }
.badge.upcoming { background: #eff6ff; border-color: #93c5fd; color: #1d4ed8; }
.muted { color: var(--muted); font-size: 14px; line-height: 1.7; }
.seat-lines { font-size: 14px; line-height: 1.85; color: #4b5563; min-height: 120px; }
.btn {
  display: inline-flex;
  align-items: center;
  justify-content: center;
  width: 100%;
  text-align: center;
  background: linear-gradient(180deg, #0f172a 0%, #0b1220 100%);
  color: white;
  border: none;
  border-radius: 16px;
  padding: 13px 18px;
  text-decoration: none;
  cursor: pointer;
  font-size: 14px;
  font-weight: 700;
  letter-spacing: .01em;
  box-shadow: 0 10px 22px rgba(15, 23, 42, 0.14);
  transition: transform .15s ease, box-shadow .15s ease, opacity .15s ease, filter .15s ease;
}
.btn:hover { transform: translateY(-2px); box-shadow: 0 14px 28px rgba(15, 23, 42, 0.18); filter: saturate(1.04); }
.btn-outline {
  background: rgba(255,255,255,0.92);
  color: var(--text);
  border: 1px solid var(--border-strong);
  box-shadow: var(--shadow-sm);
}
.btn-danger { background: linear-gradient(180deg, #ef4444 0%, #dc2626 100%); color: white; }
.btn-success { background: linear-gradient(180deg, #10b981 0%, #059669 100%); color: white; }
.section-card { padding: 24px; margin-bottom: 18px; }
.section-title { font-size: 24px; font-weight: 800; margin-bottom: 16px; letter-spacing: -0.03em; }
.notice { border-radius: 16px; padding: 16px 18px; line-height: 1.75; font-size: 14px; border: 1px solid; }
.notice.success { background: #ecfdf5; border-color: #a7f3d0; color: #065f46; }
.notice.warn { background: #fffbeb; border-color: #fcd34d; color: #92400e; }
.notice.info { background: #eff6ff; border-color: #bfdbfe; color: #1d4ed8; }
.toast { background: rgba(255,255,255,0.88); border: 1px solid var(--border); border-radius: 16px; padding: 14px 16px; margin-bottom: 18px; color: #374151; box-shadow: var(--shadow-sm); }
.form-grid { display: grid; grid-template-columns: repeat(2, 1fr); gap: 16px; }
.form-grid-1 { display: grid; grid-template-columns: 1fr; gap: 16px; }
label { display: block; font-size: 14px; font-weight: 700; margin-bottom: 8px; color: #374151; }
input[type='text'], input[type='password'], input[type='datetime-local'], input[type='date'], input[type='file'], select {
  width: 100%;
  padding: 12px 14px;
  border: 1px solid var(--border-strong);
  border-radius: 14px;
  font-size: 14px;
  outline: none;
  background: #fff;
  transition: border-color .15s ease, box-shadow .15s ease, background .15s ease;
}
input:focus, select:focus {
  border-color: #94a3b8;
  box-shadow: 0 0 0 4px rgba(148, 163, 184, 0.12);
}
.tips { font-size: 12px; color: var(--muted); margin-top: 6px; }
.divider { height: 1px; background: linear-gradient(90deg, transparent 0%, var(--border) 15%, var(--border) 85%, transparent 100%); margin: 22px 0; }
.small-box { background: var(--panel-soft); border: 1px solid var(--border); border-radius: 16px; padding: 14px 16px; font-size: 14px; color: #374151; margin-bottom: 12px; line-height: 1.75; }
.userbar { display:flex; align-items:center; gap:10px; flex-wrap:wrap; }
.table-box { overflow-x:auto; margin-top: 14px; border: 1px solid var(--border); border-radius: 18px; background:#fff; box-shadow: var(--shadow-sm); }
table { width:100%; border-collapse:collapse; font-size:14px; }
th, td { text-align:left; padding:14px 12px; border-bottom:1px solid #edf1f5; vertical-align:top; }
th { color: var(--muted); font-weight:700; background:linear-gradient(180deg, #fbfcfe 0%, #f6f8fb 100%); position: sticky; top: 0; }
.print-grid { display:grid; grid-template-columns: repeat(3, 1fr); gap:16px; }
.print-card { background:white; border:1px solid var(--border); border-radius:18px; padding:18px; text-align:center; page-break-inside: avoid; box-shadow: var(--shadow-sm); }
.print-seat-title { font-size:22px; font-weight:700; margin-bottom:10px; }
.print-qr-box { display:flex; justify-content:center; align-items:center; min-height:220px; margin-bottom:10px; }
.print-desc { font-size:13px; color:#4b5563; line-height:1.7; }
.print-url { margin-top:8px; font-size:12px; color:#6b7280; word-break:break-all; }
.print-qr-img { width:200px; height:200px; image-rendering:pixelated; }
.preset-row { display:flex; gap:10px; flex-wrap:wrap; margin-bottom:12px; }
.preset-btn { background:#fff; border:1px solid var(--border-strong); border-radius:999px; padding:8px 14px; font-size:13px; cursor:pointer; color: var(--text); box-shadow: var(--shadow-sm); }
.preset-btn:hover { background:#f9fafb; }
.toast-stack { position:fixed; top:20px; right:20px; z-index:9999; display:flex; flex-direction:column; gap:12px; }
.toast-popup { min-width:260px; max-width:360px; background:#111827; color:white; border-radius:16px; padding:14px 16px; box-shadow:0 12px 32px rgba(0,0,0,.22); opacity:0; transform:translateY(-8px); animation:toastIn .2s ease forwards, toastOut .3s ease forwards 3.2s; }
.toast-popup.success { background:#065f46; }
.toast-popup.error { background:#991b1b; }
.toast-popup.info { background:#1d4ed8; }
.toast-title { font-size:14px; font-weight:700; margin-bottom:4px; }
.toast-text { font-size:13px; line-height:1.5; }
@keyframes toastIn { to { opacity:1; transform:translateY(0); } }
@keyframes toastOut { to { opacity:0; transform:translateY(-6px); } }
.schedule-wrap { overflow-x:auto; }
.schedule-board { min-width:980px; border:1px solid var(--border); border-radius:18px; overflow:hidden; background:#fff; }
.schedule-head, .schedule-row { display:grid; grid-template-columns:110px repeat(24,1fr); gap:0; }
.schedule-head div, .schedule-row div { border-bottom:1px solid #eef2f7; }
.schedule-head .time-cell { font-size:12px; color: var(--muted); text-align:center; padding:10px 0; background:#f8fafc; }
.schedule-day { padding:12px 10px; font-size:14px; font-weight:700; border-right:1px solid #eef2f7; background:#fafbff; }
.schedule-track { position:relative; height:54px; border-right:1px solid #f1f5f9; background-image:linear-gradient(to right, #eef2f7 1px, transparent 1px); background-size:calc(100% / 24) 100%; grid-column:2 / span 24; }
.schedule-track:hover { background-color:#fafcff; }
.booking-bar { position:absolute; top:8px; height:38px; border-radius:10px; background:#f4dfb8; border:1px solid #e5c98f; color:#6b4f1d; font-size:12px; padding:6px 8px; overflow:hidden; white-space:nowrap; text-overflow:ellipsis; cursor:pointer; }
.booking-bar:hover { z-index:5; box-shadow:0 4px 12px rgba(0,0,0,.12); }
.booking-bar::after { content:attr(data-tip); position:absolute; left:50%; bottom:calc(100% + 10px); transform:translateX(-50%); min-width:220px; max-width:320px; white-space:pre-line; line-height:1.5; background:#111827; color:#fff; padding:10px 12px; border-radius:10px; font-size:12px; box-shadow:0 8px 24px rgba(0,0,0,.2); opacity:0; pointer-events:none; transition:opacity .15s ease; }
.booking-bar::before { content:''; position:absolute; left:50%; bottom:100%; transform:translateX(-50%); border:6px solid transparent; border-top-color:#111827; opacity:0; transition:opacity .15s ease; }
.booking-bar:hover::after, .booking-bar:hover::before { opacity:1; }
.booking-bar.self { background:#dbeafe; border-color:#93c5fd; color:#1d4ed8; }
.booking-bar.current { background:#f6d7a8; border-color:#d6a24d; color:#6b4f1d; font-weight:700; }
.booking-bar.self.current { background:#bfdbfe; border-color:#60a5fa; color:#1d4ed8; font-weight:700; }
.legend-row { display:flex; gap:16px; flex-wrap:wrap; align-items:center; margin-bottom:12px; font-size:13px; color:#4b5563; }
.legend-item { display:flex; align-items:center; gap:8px; }
.legend-dot { width:18px; height:12px; border-radius:6px; border:1px solid #d1d5db; display:inline-block; }
.legend-dot.mine { background:#dbeafe; border-color:#93c5fd; }
.legend-dot.other { background:#f4dfb8; border-color:#e5c98f; }
.legend-dot.current { background:#f6d7a8; border-color:#d6a24d; }
.legend-line { width:18px; height:0; border-top:2px solid #ef4444; display:inline-block; }
.now-line { position:absolute; top:0; bottom:0; width:2px; background:#ef4444; z-index:2; }
.now-label { position:absolute; top:-18px; transform:translateX(-50%); color:#ef4444; font-size:12px; white-space:nowrap; }
.inline-form { display:flex; gap:8px; flex-wrap:wrap; align-items:center; }
.admin-shell { display:flex; min-height:100vh; background:transparent; }
.admin-sidebar { width:260px; background:rgba(255,255,255,0.92); backdrop-filter:blur(10px); border-right:1px solid rgba(229,231,235,0.9); padding:24px 18px; box-shadow: var(--shadow-sm); }
.admin-sidebar-title { font-size:20px; font-weight:700; margin-bottom:8px; letter-spacing:-0.02em; }
.admin-sidebar-sub { font-size:13px; color: var(--muted); margin-bottom:20px; }
.admin-nav { display:flex; flex-direction:column; gap:8px; }
.admin-nav-link { display:block; padding:13px 15px; border-radius:14px; text-decoration:none; color: var(--text); border:1px solid transparent; transition: all .15s ease; font-weight: 600; }
.admin-nav-link:hover { background:#f9fafb; border-color: var(--border); }
.admin-nav-link.active { background:linear-gradient(180deg, #0f172a 0%, #111827 100%); color:#fff; box-shadow: 0 10px 22px rgba(17, 24, 39, 0.14); }
.admin-main { flex:1; padding:30px 26px 42px; }
.admin-page { max-width:none; margin:0; padding:0; }
@media print {
  .no-print { display:none !important; }
  body { background:white; }
  .container { max-width:none; padding:0; }
  .print-grid { grid-template-columns:repeat(3,1fr); gap:12px; }
  .print-card { box-shadow:none; }
}
@media (max-width:960px) { .seat-grid { grid-template-columns:repeat(2,1fr); } .print-grid { grid-template-columns:repeat(2,1fr); } }
@media (max-width:900px) { .admin-shell { flex-direction:column; } .admin-sidebar { width:100%; border-right:none; border-bottom:1px solid var(--border); } }
@media (max-width:720px) { .stats, .form-grid, .seat-grid, .print-grid { grid-template-columns:1fr; } .title { font-size:24px; } }
</style>
"""


def toast_script_block():
    return """
    <div id='toast-stack' class='toast-stack'></div>
    <script>
    window.addEventListener('DOMContentLoaded', function(){
        const msg = window.__popupMessage || '';
        const level = window.__popupLevel || 'info';
        if (msg) {
            const stack = document.getElementById('toast-stack');
            if (stack) {
                const item = document.createElement('div');
                item.className = 'toast-popup ' + level;
                item.innerHTML = '<div class="toast-title">系统提示</div><div class="toast-text"></div>';
                item.querySelector('.toast-text').textContent = msg;
                stack.appendChild(item);
                setTimeout(() => item.remove(), 3800);
            }
        }
    });
    </script>
    """

LOGIN_TEMPLATE = """
<!DOCTYPE html><html lang='zh-CN'><head><meta charset='UTF-8'><meta name='viewport' content='width=device-width, initial-scale=1.0'><title>登录</title>{{ css|safe }}</head>
<body><div class='auth-container'><div class='card section-card'><div class='pill'>工位预约系统</div><div class='title'>卡旺卡工位预约管理Demo</div><div class='subtitle'>欢迎体验卡旺卡工位预约管理系统。释放工位时必须当前登录人就是占用人本人。</div>{% if message %}<div class='notice' style='margin-top:18px; background:#fef2f2; border-color:#fecaca; color:#b91c1c;'>{{ message }}</div>{% endif %}<form method='post' style='margin-top:18px;'><div class='form-grid-1'><div><label>工号</label><input type='text' name='employee_no' placeholder='请输入工号，例如：admin 或 1001'></div><div><label>密码</label><input type='password' name='password' placeholder='请输入密码'></div></div><div class='divider'></div><button class='btn' type='submit'>登录</button></form><div class='small-box' style='margin-top:18px;'><strong>内置测试账号：</strong><br>管理员：admin / admin123 / IT部<br>张三：1001 / 123456 / 营运部<br>李四：1002 / 123456 / 财务部<br>王五：1003 / 123456 / IT部</div></div></div></body></html>
"""

INDEX_TEMPLATE = """
<!DOCTYPE html><html lang='zh-CN'><head><meta charset='UTF-8'><meta name='viewport' content='width=device-width, initial-scale=1.0'><title>首页</title>{{ css|safe }}
<style>
.home-search-card { margin-bottom: 22px; background: linear-gradient(135deg, rgba(255,255,255,0.98) 0%, rgba(247,250,255,0.98) 100%); border: 1px solid rgba(226,232,240,0.95); }
.home-search-actions { display:flex; gap:12px; align-items:flex-end; margin-top: 8px; }
.home-seat-grid { display:grid; grid-template-columns: repeat(auto-fill, minmax(260px, 1fr)); gap:18px; }
.home-seat-card { min-height: 310px; padding: 22px; border-radius: 20px; box-shadow: 0 1px 2px rgba(0,0,0,.04); display:flex; flex-direction:column; }
.home-seat-card:hover { transform: none; box-shadow: 0 1px 2px rgba(0,0,0,.04); border-color: rgba(229,231,235,0.9); }
.home-seat-card .seat-lines { flex: 1; min-height: 120px; }
.home-seat-card .btn { border-radius: 16px; padding: 13px 16px; box-shadow: 0 8px 18px rgba(17, 24, 39, 0.12); }
.home-seat-title { font-size: 26px; line-height: 1.12; letter-spacing: -0.03em; word-break: break-word; }
</style></head>
<body><div class='container'>
<div class='topbar'>
  <div>
    <div class='pill'>工位预约首页</div>
    <div class='title'>卡旺卡共享工位预约Demo</div>
    <div class='subtitle'>支持搜索工位、搜索伙伴姓名，并快速判断当前占用与后续预约状态。</div>
  </div>
  <div class='userbar'>
    <div class='pill'>当前登录：{{ current_user.name }}（{{ current_user.employee_no }} / {{ current_user.department or '未设置部门' }}）</div>
    {% if current_user.role == 'admin' %}<a class='btn btn-outline' style='width:auto;' href='{{ url_for("admin_home") }}'>后台管理</a>{% endif %}
    <a class='btn btn-outline' style='width:auto;' href='{{ url_for("my_records") }}'>我的记录</a>
    <a class='btn btn-danger' style='width:auto;' href='{{ url_for("logout") }}'>退出登录</a>
  </div>
</div>

{% if message %}<div class='toast'>{{ message }}</div>{% endif %}

<div class='card section-card home-search-card'>
  <div class='section-title'>工位搜索 / 找人</div>
  <form method='get'>
    <div class='form-grid-1'>
      <div>
        <label>输入工位名称或伙伴姓名</label>
        <input type='text' name='q' value='{{ q }}' placeholder='例如：靠窗财务工位1 / 李四'>
        <div class='tips'>支持按工位名称搜索，也支持按伙伴姓名查找当前坐席。</div>
      </div>
    </div>
    <div class='home-search-actions'>
      <button class='btn' type='submit' style='width:auto;'>搜索</button>
      {% if q %}<a class='btn btn-outline' href='{{ url_for("index") }}' style='width:auto;'>清空</a>{% endif %}
    </div>
  </form>

  {% if q %}
    <div class='divider'></div>
    {% if people_matches %}
      <div class='notice success' style='margin-bottom:12px;'>
        <div><strong>找人结果</strong></div>
        {% for item in people_matches %}
          <div style='margin-top:6px;'>{{ item.user_name }} 当前在：<strong>{{ item.seat_display }}</strong>，结束时间：{{ item.use_end_at_display }}</div>
        {% endfor %}
      </div>
    {% elif filtered_count > 0 %}
      <div class='small-box'>共找到 {{ filtered_count }} 个匹配的结果。</div>
    {% else %}
      <div class='small-box'>当前没有找到“{{ q }}”正在使用的工位。</div>
    {% endif %}
  {% endif %}
</div>

<div class='stats'>
  <div class='card stat-card'><div class='stat-label'>工位总数</div><div class='stat-value'>{{ total }}</div></div>
  <div class='card stat-card'><div class='stat-label'>当前已占用</div><div class='stat-value'>{{ occupied }}</div></div>
  <div class='card stat-card'><div class='stat-label'>当前可使用</div><div class='stat-value'>{{ free }}</div></div>
</div>

<div class='home-seat-grid'>
{% for seat in seats %}
  <div class='card seat-card home-seat-card'>
    <div class='seat-head'>
      <div class='seat-title home-seat-title'>{{ seat.display_title }}</div>
      {% if seat.current_item %}<div class='badge busy'>已占用</div>{% elif seat.next_item %}<div class='badge upcoming'>后续有预约</div>{% else %}<div class='badge free'>空闲</div>{% endif %}
    </div>
    <div class='seat-lines'>
      {% if seat.current_item %}
        <div>占用人：{{ seat.current_item.user_name }}</div>
        <div>工号：{{ seat.current_item.employee_no }}</div>
        <div>开始：{{ seat.current_item.start_at_display }}</div>
        <div>结束：{{ seat.current_item.use_end_at_display }}</div>
      {% elif seat.next_item %}
        <div class='muted'>当前可使用</div>
        <div>后续预约：{{ seat.next_item.user_name }}</div>
        <div>开始：{{ seat.next_item.start_at_display }}</div>
        <div>结束：{{ seat.next_item.use_end_at_display }}</div>
      {% else %}
        <div class='muted'>当前空闲，且暂无后续预约冲突。</div>
      {% endif %}
    </div>
    <div style='margin-top:auto;'><a class='btn' href='{{ url_for("seat_page", seat_id=seat.id) }}'>打开工位页面</a></div>
  </div>
{% endfor %}
</div>
</div></body></html>
"""

ADMIN_HOME_TEMPLATE = """
<!DOCTYPE html><html lang='zh-CN'><head><meta charset='UTF-8'><meta name='viewport' content='width=device-width, initial-scale=1.0'><title>后台管理</title>{{ css|safe }}</head>
<body>
<div class='admin-shell'>
    <aside class='admin-sidebar'>
        <div class='admin-sidebar-title'>后台管理</div>
        <div class='admin-sidebar-sub'>左侧菜单切换，右侧查看与操作</div>
        <nav class='admin-nav'>
            <a class='admin-nav-link active' href='{{ url_for("admin_home") }}'>管理首页</a>
            <a class='admin-nav-link' href='{{ url_for("admin_users") }}'>用户管理</a>
            <a class='admin-nav-link' href='{{ url_for("admin_seats") }}'>工位管理</a>
            <a class='admin-nav-link' href='{{ url_for("admin_reservations") }}'>预约管理</a>
            <a class='admin-nav-link' href='{{ url_for("admin_reports") }}'>统计报表</a>
            <a class='admin-nav-link' href='{{ url_for("qr_print_page") }}' target='_blank'>二维码打印</a>
            <a class='admin-nav-link' href='{{ url_for("index") }}'>返回首页</a>
        </nav>
    </aside>
    <main class='admin-main'>
        <div class='container admin-page'>
            <div class='page-head'>
                <div>
                    <div class='pill'>管理员后台</div>
                    <div class='title' style='margin-top:8px;'>后台管理首页</div>
                    <div class='subtitle'>点击左侧菜单，右侧进入对应管理页面。</div>
                </div>
            </div>
            <div class='seat-grid'>
                <div class='card seat-card'><div class='seat-title' style='margin-bottom:12px;'>用户管理</div><div class='muted'>创建、编辑、删除、禁用、重置密码、Excel导入。</div><div style='margin-top:16px;'><a class='btn' href='{{ url_for("admin_users") }}'>进入</a></div></div>
                <div class='card seat-card'><div class='seat-title' style='margin-bottom:12px;'>工位管理</div><div class='muted'>创建工位、命名工位、维护工位信息。</div><div style='margin-top:16px;'><a class='btn' href='{{ url_for("admin_seats") }}'>进入</a></div></div>
                <div class='card seat-card'><div class='seat-title' style='margin-bottom:12px;'>预约管理</div><div class='muted'>查看所有预约，并支持强制取消、强制释放。</div><div style='margin-top:16px;'><a class='btn' href='{{ url_for("admin_reservations") }}'>进入</a></div></div>
                <div class='card seat-card'><div class='seat-title' style='margin-bottom:12px;'>统计报表</div><div class='muted'>按部门、按人、按工位查看使用率和预约数据。</div><div style='margin-top:16px;'><a class='btn' href='{{ url_for("admin_reports") }}'>进入</a></div></div>
                <div class='card seat-card'><div class='seat-title' style='margin-bottom:12px;'>二维码打印页</div><div class='muted'>批量打印工位二维码，贴到对应工位现场。</div><div style='margin-top:16px;'><a class='btn' href='{{ url_for("qr_print_page") }}' target='_blank'>进入</a></div></div>
            </div>
        </div>
    </main>
</div>
</body></html>
"""

ADMIN_USERS_TEMPLATE = """
<!DOCTYPE html><html lang='zh-CN'><head><meta charset='UTF-8'><meta name='viewport' content='width=device-width, initial-scale=1.0'><title>用户管理</title>{{ css|safe }}</head>
<body><script>window.__popupMessage={{ popup_message|tojson }}; window.__popupLevel={{ popup_level|tojson }};</script>{{ toast_block|safe }}
<div class='admin-shell'>
    <aside class='admin-sidebar'>
        <div class='admin-sidebar-title'>后台管理</div>
        <div class='admin-sidebar-sub'>左侧菜单切换，右侧查看与操作</div>
        <nav class='admin-nav'>
            <a class='admin-nav-link' href='{{ url_for("admin_home") }}'>管理首页</a>
            <a class='admin-nav-link active' href='{{ url_for("admin_users") }}'>用户管理</a>
            <a class='admin-nav-link' href='{{ url_for("admin_seats") }}'>工位管理</a>
            <a class='admin-nav-link' href='{{ url_for("admin_reservations") }}'>预约管理</a>
            <a class='admin-nav-link' href='{{ url_for("admin_reports") }}'>统计报表</a>
            <a class='admin-nav-link' href='{{ url_for("qr_print_page") }}' target='_blank'>二维码打印</a>
            <a class='admin-nav-link' href='{{ url_for("index") }}'>返回首页</a>
        </nav>
    </aside>
    <main class='admin-main'><div class='container admin-page'>
        <div class='page-head'><div><div class='pill'>管理员后台</div><div class='title' style='margin-top:8px;'>用户管理</div><div class='subtitle'>管理员可创建、编辑、删除、禁用账号，并支持 Excel 批量导入用户。</div></div></div>
        <div class='card section-card'><div class='section-title'>筛选条件</div><form method='get' class='form-grid'><div><label>姓名</label><input type='text' name='name' value='{{ filters.name }}' placeholder='例如：张三'></div><div><label>部门</label><input type='text' name='department' value='{{ filters.department }}' placeholder='例如：IT部'></div><div><label>角色</label><select name='role'><option value=''>全部</option><option value='employee' {% if filters.role=='employee' %}selected{% endif %}>普通员工</option><option value='admin' {% if filters.role=='admin' %}selected{% endif %}>管理员</option></select></div><div><label>状态</label><select name='status'><option value=''>全部</option><option value='active' {% if filters.status=='active' %}selected{% endif %}>启用</option><option value='disabled' {% if filters.status=='disabled' %}selected{% endif %}>禁用</option></select></div><div style='display:flex; align-items:end; gap:12px;'><button class='btn' type='submit'>筛选</button><a class='btn btn-outline' href='{{ url_for("admin_users") }}'>清空</a></div></form></div>
        <div class='form-grid'>
            <div class='card section-card'><div class='section-title'>新增单个用户</div><form method='post' action='{{ url_for("admin_create_user") }}'><div class='form-grid-1'><div><label>工号</label><input type='text' name='employee_no' required></div><div><label>姓名</label><input type='text' name='name' required></div><div><label>部门</label><input type='text' name='department'></div><div><label>密码</label><input type='text' name='password' required></div><div><label>角色</label><select name='role'><option value='employee'>普通员工</option><option value='admin'>管理员</option></select></div></div><div class='divider'></div><button class='btn btn-success' type='submit'>创建用户</button></form></div>
            <div class='card section-card'><div class='section-title'>Excel 批量导入</div><form method='post' action='{{ url_for("admin_import_users") }}' enctype='multipart/form-data'><div class='form-grid-1'><div><label>上传 Excel 文件</label><input type='file' name='file' accept='.xlsx' required></div></div><div class='small-box' style='margin-top:16px;'><strong>Excel 列头格式：</strong><br>工号 / 姓名 / 部门 / 密码 / 角色</div><div style='display:flex; gap:12px; flex-wrap:wrap; margin-top:16px;'><a class='btn btn-outline' style='width:auto;' href='{{ url_for("admin_user_template") }}'>下载导入模板</a></div><div class='divider'></div><button class='btn' type='submit'>导入用户</button></form></div>
        </div>
        <div class='card section-card'><div class='section-title'>当前用户列表</div><div class='table-box'><table><thead><tr><th>工号</th><th>姓名</th><th>部门</th><th>角色</th><th>状态</th><th>创建时间</th><th>操作</th></tr></thead><tbody>{% for user in users %}<tr><td>{{ user.employee_no }}</td><td>{{ user.name }}</td><td>{{ user.department or '未设置' }}</td><td>{{ '管理员' if user.role == 'admin' else '普通员工' }}</td><td>{{ user.status }}</td><td>{{ user.created_at_display }}</td><td><div style='display:flex; flex-direction:column; gap:10px; min-width:260px;'><form method='post' action='{{ url_for("admin_update_user", employee_no=user.employee_no) }}' class='inline-form'><input type='text' name='name' value='{{ user.name }}' style='width:80px; padding:8px 10px; border-radius:10px;' required><input type='text' name='department' value='{{ user.department or "" }}' style='width:90px; padding:8px 10px; border-radius:10px;'><select name='role' style='width:90px; padding:8px 10px; border-radius:10px;'><option value='employee' {% if user.role == 'employee' %}selected{% endif %}>普通员工</option><option value='admin' {% if user.role == 'admin' %}selected{% endif %}>管理员</option></select><button class='btn btn-outline' style='width:auto; padding:8px 12px;'>保存编辑</button></form><div class='inline-form'>{% if user.employee_no != 'admin' %}<form method='post' action='{{ url_for("admin_toggle_user_status", employee_no=user.employee_no) }}' style='margin:0;'>{% if user.status == 'active' %}<button class='btn btn-outline' style='width:auto; padding:8px 12px;'>禁用用户</button>{% else %}<button class='btn btn-success' style='width:auto; padding:8px 12px;'>启用用户</button>{% endif %}</form>{% endif %}<form method='post' action='{{ url_for("admin_reset_password", employee_no=user.employee_no) }}' style='margin:0;' class='inline-form'><input type='text' name='new_password' placeholder='新密码' style='width:120px; padding:8px 10px; border-radius:10px;' required><button class='btn btn-outline' style='width:auto; padding:8px 12px;'>重置密码</button></form>{% if user.employee_no != 'admin' %}<form method='post' action='{{ url_for("admin_delete_user", employee_no=user.employee_no) }}' style='margin:0;' onsubmit='return confirm("确认删除用户 {{ user.employee_no }} 吗？")'><button class='btn btn-danger' style='width:auto; padding:8px 12px;'>删除用户</button></form>{% endif %}</div></div></td></tr>{% endfor %}{% if not users %}<tr><td colspan='7'>暂无用户</td></tr>{% endif %}</tbody></table></div></div>
    </div></main>
</div>
</body></html>
"""

ADMIN_SEATS_TEMPLATE = """
<!DOCTYPE html><html lang='zh-CN'><head><meta charset='UTF-8'><meta name='viewport' content='width=device-width, initial-scale=1.0'><title>工位管理</title>{{ css|safe }}</head>
<body><script>window.__popupMessage={{ popup_message|tojson }}; window.__popupLevel={{ popup_level|tojson }};</script>{{ toast_block|safe }}
<div class='admin-shell'>
    <aside class='admin-sidebar'>
        <div class='admin-sidebar-title'>后台管理</div>
        <div class='admin-sidebar-sub'>左侧菜单切换，右侧查看与操作</div>
        <nav class='admin-nav'>
            <a class='admin-nav-link' href='{{ url_for("admin_home") }}'>管理首页</a>
            <a class='admin-nav-link' href='{{ url_for("admin_users") }}'>用户管理</a>
            <a class='admin-nav-link active' href='{{ url_for("admin_seats") }}'>工位管理</a>
            <a class='admin-nav-link' href='{{ url_for("admin_reservations") }}'>预约管理</a>
            <a class='admin-nav-link' href='{{ url_for("admin_reports") }}'>统计报表</a>
            <a class='admin-nav-link' href='{{ url_for("qr_print_page") }}' target='_blank'>二维码打印</a>
            <a class='admin-nav-link' href='{{ url_for("index") }}'>返回首页</a>
        </nav>
    </aside>
    <main class='admin-main'><div class='container admin-page'>
        <div class='page-head'><div><div class='pill'>管理员后台</div><div class='title' style='margin-top:8px;'>工位管理</div><div class='subtitle'>创建工位、命名工位，并维护工位信息。</div></div></div>
        <div class='card section-card'><div class='section-title'>创建工位</div><form method='post' action='{{ url_for("admin_create_seat") }}' class='form-grid'><div><label>工位名称</label><input type='text' name='seat_name' required></div><div style='display:flex; align-items:end;'><button class='btn btn-success' type='submit'>创建工位</button></div></form></div>
        <div class='card section-card'><div class='section-title'>当前工位列表</div><div class='table-box'><table><thead><tr><th>工位名称</th><th>状态</th><th>创建时间</th><th>操作</th></tr></thead><tbody>{% for seat in seats %}<tr><td>{{ seat.seat_name or '未命名' }}</td><td>{{ seat.status }}</td><td>{{ seat.created_at_display }}</td><td><form method='post' action='{{ url_for("admin_update_seat", seat_code=seat.seat_code) }}' class='inline-form'><input type='text' name='seat_name' value='{{ seat.seat_name or "" }}' placeholder='工位名称' style='width:180px; padding:8px 10px; border-radius:10px;'><button class='btn btn-outline' style='width:auto; padding:8px 12px;'>保存名称</button></form></td></tr>{% endfor %}{% if not seats %}<tr><td colspan='4'>暂无工位</td></tr>{% endif %}</tbody></table></div></div>
    </div></main>
</div>
</body></html>
"""

ADMIN_RESERVATIONS_TEMPLATE = """
<!DOCTYPE html><html lang='zh-CN'><head><meta charset='UTF-8'><meta name='viewport' content='width=device-width, initial-scale=1.0'><title>预约管理</title>{{ css|safe }}</head>
<body><script>window.__popupMessage={{ popup_message|tojson }}; window.__popupLevel={{ popup_level|tojson }};</script>{{ toast_block|safe }}
<div class='admin-shell'>
    <aside class='admin-sidebar'>
        <div class='admin-sidebar-title'>后台管理</div>
        <div class='admin-sidebar-sub'>左侧菜单切换，右侧查看与操作</div>
        <nav class='admin-nav'>
            <a class='admin-nav-link' href='{{ url_for("admin_home") }}'>管理首页</a>
            <a class='admin-nav-link' href='{{ url_for("admin_users") }}'>用户管理</a>
            <a class='admin-nav-link' href='{{ url_for("admin_seats") }}'>工位管理</a>
            <a class='admin-nav-link active' href='{{ url_for("admin_reservations") }}'>预约管理</a>
            <a class='admin-nav-link' href='{{ url_for("admin_reports") }}'>统计报表</a>
            <a class='admin-nav-link' href='{{ url_for("qr_print_page") }}' target='_blank'>二维码打印</a>
            <a class='admin-nav-link' href='{{ url_for("index") }}'>返回首页</a>
        </nav>
    </aside>
    <main class='admin-main'><div class='container admin-page'>
        <div class='page-head'><div><div class='pill'>管理员后台</div><div class='title' style='margin-top:8px;'>预约管理</div><div class='subtitle'>查看所有预约记录，并可强制取消未来预约或强制释放当前占用。</div></div></div>
        <div class='card section-card'><div class='section-title'>筛选条件</div><form method='get' class='form-grid'><div><label>开始日期</label><input type='date' name='start_date' value='{{ filters.start_date }}'></div><div><label>结束日期</label><input type='date' name='end_date' value='{{ filters.end_date }}'></div><div><label>姓名</label><input type='text' name='name' placeholder='例如：张三' value='{{ filters.name }}'></div><div style='display:flex; align-items:end;'><button class='btn' type='submit'>筛选</button></div></form></div>
        <div class='card section-card'><div class='section-title'>全部预约</div><div class='table-box'><table><thead><tr><th>工位</th><th>预约人</th><th>部门</th><th>开始</th><th>结束</th><th>状态</th><th>操作</th></tr></thead><tbody>{% for row in rows %}<tr><td>{{ row.seat_display }}</td><td>{{ row.user_name }}（{{ row.employee_no }}）</td><td>{{ row.department or '未设置' }}</td><td>{{ row.start_at_display }}</td><td>{{ row.use_end_at_display }}</td><td>{{ row.status_label }}</td><td>{% if row.can_force_release %}<form method='post' action='{{ url_for("admin_manage_reservation", reservation_id=row.id) }}' style='margin:0;'><input type='hidden' name='action' value='force_release'><button class='btn btn-danger' style='width:auto; padding:8px 12px;'>强制释放</button></form>{% elif row.can_force_cancel %}<form method='post' action='{{ url_for("admin_manage_reservation", reservation_id=row.id) }}' style='margin:0;'><input type='hidden' name='action' value='force_cancel'><button class='btn btn-outline' style='width:auto; padding:8px 12px;'>强制取消</button></form>{% else %}-{% endif %}</td></tr>{% endfor %}{% if not rows %}<tr><td colspan='7'>暂无预约</td></tr>{% endif %}</tbody></table></div></div>
    </div></main>
</div>
</body></html>
"""

ADMIN_REPORTS_TEMPLATE = """
<!DOCTYPE html><html lang='zh-CN'><head><meta charset='UTF-8'><meta name='viewport' content='width=device-width, initial-scale=1.0'><title>统计报表</title>{{ css|safe }}</head>
<body>
<div class='admin-shell'>
    <aside class='admin-sidebar'>
        <div class='admin-sidebar-title'>后台管理</div>
        <div class='admin-sidebar-sub'>左侧菜单切换，右侧查看与操作</div>
        <nav class='admin-nav'>
            <a class='admin-nav-link' href='{{ url_for("admin_home") }}'>管理首页</a>
            <a class='admin-nav-link' href='{{ url_for("admin_users") }}'>用户管理</a>
            <a class='admin-nav-link' href='{{ url_for("admin_seats") }}'>工位管理</a>
            <a class='admin-nav-link' href='{{ url_for("admin_reservations") }}'>预约管理</a>
            <a class='admin-nav-link active' href='{{ url_for("admin_reports") }}'>统计报表</a>
            <a class='admin-nav-link' href='{{ url_for("qr_print_page") }}' target='_blank'>二维码打印</a>
            <a class='admin-nav-link' href='{{ url_for("index") }}'>返回首页</a>
        </nav>
    </aside>
    <main class='admin-main'><div class='container admin-page'>
        <div class='page-head'><div><div class='pill'>管理员后台</div><div class='title' style='margin-top:8px;'>预约统计报表</div><div class='subtitle'>按部门、按人、按工位查看预约时长与使用率。</div></div></div>
        <div class='card section-card'><div class='section-title'>筛选条件</div><form method='get' class='form-grid'><div><label>开始日期</label><input type='date' name='start_date' value='{{ filters.start_date }}'></div><div><label>结束日期</label><input type='date' name='end_date' value='{{ filters.end_date }}'></div><div><label>姓名</label><input type='text' name='name' placeholder='例如：张三' value='{{ filters.name }}'></div><div><label>部门</label><input type='text' name='department' placeholder='例如：IT部' value='{{ filters.department }}'></div><div style='display:flex; align-items:end;'><button class='btn' type='submit'>筛选</button></div></form></div>
        <div class='stats'><div class='card stat-card'><div class='stat-label'>总预约次数</div><div class='stat-value'>{{ summary.total_reservations }}</div></div><div class='card stat-card'><div class='stat-label'>总预约小时数</div><div class='stat-value'>{{ summary.total_hours }}</div></div><div class='card stat-card'><div class='stat-label'>工位平均使用率</div><div class='stat-value'>{{ summary.avg_utilization }}%</div></div></div>
        <div class='card section-card'><div class='section-title'>按部门统计</div><div class='table-box'><table><thead><tr><th>部门</th><th>预约次数</th><th>总小时数</th><th>占比</th></tr></thead><tbody>{% for row in by_department %}<tr><td>{{ row.name }}</td><td>{{ row.count }}</td><td>{{ row.hours }}</td><td>{{ row.ratio }}%</td></tr>{% endfor %}{% if not by_department %}<tr><td colspan='4'>暂无数据</td></tr>{% endif %}</tbody></table></div></div>
        <div class='card section-card'><div class='section-title'>按人统计</div><div class='table-box'><table><thead><tr><th>工号</th><th>姓名</th><th>部门</th><th>预约次数</th><th>总小时数</th></tr></thead><tbody>{% for row in by_user %}<tr><td>{{ row.employee_no }}</td><td>{{ row.name }}</td><td>{{ row.department }}</td><td>{{ row.count }}</td><td>{{ row.hours }}</td></tr>{% endfor %}{% if not by_user %}<tr><td colspan='5'>暂无数据</td></tr>{% endif %}</tbody></table></div></div>
        <div class='card section-card'><div class='section-title'>按工位统计</div><div class='table-box'><table><thead><tr><th>工位</th><th>预约次数</th><th>总小时数</th><th>使用率</th></tr></thead><tbody>{% for row in by_seat %}<tr><td>{{ row.seat_display }}</td><td>{{ row.count }}</td><td>{{ row.hours }}</td><td>{{ row.utilization }}%</td></tr>{% endfor %}{% if not by_seat %}<tr><td colspan='4'>暂无数据</td></tr>{% endif %}</tbody></table></div></div>
    </div></main>
</div>
</body></html>
"""

QR_PRINT_TEMPLATE = """
<!DOCTYPE html><html lang='zh-CN'><head><meta charset='UTF-8'><meta name='viewport' content='width=device-width, initial-scale=1.0'><title>二维码打印页</title>{{ css|safe }}</head>
<body><div class='container'><div class='page-head no-print'><div><div class='pill'>工位二维码打印页</div><div class='title' style='margin-top:8px;'>打印后贴到对应工位</div><div class='subtitle'>每张二维码都对应一个独立工位。伙伴扫码后会直接进入对应工位预约页面。</div></div><div style='display:flex; gap:12px; flex-wrap:wrap;'><a class='btn btn-outline' style='width:auto;' href='{{ url_for("admin_home") }}'>返回后台首页</a><button class='btn' style='width:auto;' onclick='window.print()'>打印当前页面</button></div></div><div class='print-grid'>{% for seat in seats %}<div class='print-card'><div class='print-seat-title'>{{ seat.name or '未命名工位' }}</div><div class='print-qr-box'><img class='print-qr-img' src='{{ seat.qr_data_uri }}' alt='二维码'></div><div class='print-desc'>扫码预约 / 本人登录释放</div><div class='print-url'>{{ seat.url }}</div></div>{% endfor %}</div></div></body></html>
"""

SEAT_TEMPLATE = """
<!DOCTYPE html><html lang='zh-CN'><head><meta charset='UTF-8'><meta name='viewport' content='width=device-width, initial-scale=1.0'><title>{{ seat_display_title }}</title>{{ css|safe }}
<script>
function pad(n){return String(n).padStart(2,'0');}
function formatLocal(dt){return dt.getFullYear()+'-'+pad(dt.getMonth()+1)+'-'+pad(dt.getDate())+'T'+pad(dt.getHours())+':'+pad(dt.getMinutes());}
function getStartInput(){return document.getElementById('start_at');}
function getEndInput(){return document.getElementById('use_end_at');}
function setEndFromStart(minutes){const s=getStartInput(), e=getEndInput(); if(!s||!e||!s.value) return; const start=new Date(s.value); const end=new Date(start.getTime()+minutes*60000); const midnight=new Date(start); midnight.setHours(24,0,0,0); e.value=formatLocal(end>midnight?midnight:end);} 
function setOffWork(){const s=getStartInput(), e=getEndInput(); if(!s||!e||!s.value) return; const start=new Date(s.value); const t=new Date(start); t.setHours(18,0,0,0); e.value=formatLocal(t);} 
function setMidnight(){const s=getStartInput(), e=getEndInput(); if(!s||!e||!s.value) return; const start=new Date(s.value); const t=new Date(start); t.setHours(24,0,0,0); e.value=formatLocal(t);} 
function alignEndAfterStart(){const s=getStartInput(), e=getEndInput(); if(!s||!e||!s.value) return; if(!e.value){setEndFromStart(120); return;} const start=new Date(s.value); const end=new Date(e.value); if(end<=start){setEndFromStart(120);} }
window.__popupMessage={{ popup_message|tojson }}; window.__popupLevel={{ popup_level|tojson }};
</script></head>
<body>{{ toast_block|safe }}<div class='container' style='max-width:1080px;'><div class='page-head'><div><div class='muted'>扫码进入</div><div class='title' style='margin-top:4px;'>{{ seat_display_title }}</div></div><div style='display:flex; gap:12px; flex-wrap:wrap;'><a class='btn btn-outline' style='width:auto;' href='{{ url_for("index") }}'>返回首页</a><a class='btn btn-danger' style='width:auto;' href='{{ url_for("logout") }}'>退出登录</a></div></div><div class='toast'>当前登录：{{ current_user.name }}（{{ current_user.employee_no }} / {{ current_user.department or '未设置部门' }}）</div>{% if message %}<div class='toast'>{{ message }}</div>{% endif %}<div class='card section-card'><div class='section-title'>未来7天预约状态</div><div class='legend-row'><div class='legend-item'><span class='legend-dot mine'></span><span>我的预约</span></div><div class='legend-item'><span class='legend-dot other'></span><span>他人预约</span></div><div class='legend-item'><span class='legend-dot current'></span><span>当前进行中</span></div><div class='legend-item'><span class='legend-line'></span><span>当前时间</span></div></div><div class='schedule-wrap'><div class='schedule-board'><div class='schedule-head'><div class='schedule-day'>日期</div>{% for hour in hours %}<div class='time-cell'>{{ hour }}</div>{% endfor %}</div>{% for day in schedule_days %}<div class='schedule-row'><div class='schedule-day'>{{ day.label }}</div><div class='schedule-track'>{% if day.is_today %}<div class='now-line' style='left: {{ day.now_left }}%;'><div class='now-label'>{{ day.now_label }}</div></div>{% endif %}{% for item in day.bookings %}<div class='booking-bar {% if item.is_self %}self{% endif %} {% if item.is_current %}current{% endif %}' style='left: {{ item.left }}%; width: {{ item.width }}%;' title='{{ item.title }}' data-tip='{{ item.tip }}'>{{ item.text }}</div>{% endfor %}</div></div>{% endfor %}</div></div></div><div class='card section-card'><div class='section-title'>最近相关预约</div>{% if reservation %}<div class='notice warn'><div><strong>{% if reservation.phase == 'occupied' %}该工位当前已被占用{% else %}该工位已有后续预约{% endif %}</strong></div><div>预约人：{{ reservation.user_name }}（{{ reservation.employee_no }}）</div><div>开始：{{ reservation.start_at_display }}</div><div>结束：{{ reservation.use_end_at_display }}</div><div>系统最晚自动释放：该预约开始当天 24:00</div></div>{% else %}<div class='notice success'><div><strong>该工位当前没有冲突预约</strong></div><div>当前登录人可预约一个开始时间和结束时间，且结束时间必须在开始当天 24:00 前。</div></div>{% endif %}</div><div class='card section-card'><div class='section-title'>预约操作区</div><form method='post'><div class='form-grid'><div><label>当前登录伙伴</label><input type='text' value='{{ current_user.name }}（{{ current_user.employee_no }} / {{ current_user.department or "未设置部门" }}）' disabled></div><div><label>预约开始时间</label><input id='start_at' type='datetime-local' name='start_at' value='{{ default_start_at }}' onchange='alignEndAfterStart()'><div class='tips'>可预约未来时间，例如提前两天预订。</div></div></div><div class='form-grid' style='margin-top:16px;'><div style='grid-column:1 / -1;'><label>使用到几点</label><div class='preset-row'><button type='button' class='preset-btn' onclick='setEndFromStart(30)'>30分钟</button><button type='button' class='preset-btn' onclick='setEndFromStart(60)'>1小时</button><button type='button' class='preset-btn' onclick='setEndFromStart(120)'>2小时</button><button type='button' class='preset-btn' onclick='setOffWork()'>下班前（18:00）</button><button type='button' class='preset-btn' onclick='setMidnight()'>24点前</button></div><input id='use_end_at' type='datetime-local' name='use_end_at' value='{{ default_use_end_at }}'><div class='tips'>默认结束时间 = 开始时间 + 2小时。你也可以选择当天任意时间，但最晚不能超过开始当天24:00。</div></div></div><div class='divider'></div><button type='submit' name='action' value='reserve' class='btn'>以当前登录身份预约该工位</button>{% if reservation and reservation.user_id == current_user.id and reservation.phase == 'occupied' %}<div class='small-box' style='margin-top:12px;'>当前这条预约已开始，且由你本人创建，可直接释放。</div><button type='submit' name='action' value='release' class='btn btn-danger'>本人释放当前工位</button>{% elif reservation and reservation.user_id == current_user.id and reservation.phase == 'upcoming' %}<div class='small-box' style='margin-top:12px;'>这是一条未开始的预约。你可前往“我的记录”中取消该预约。</div>{% elif reservation and reservation.user_id != current_user.id %}<div class='notice info' style='margin-top:12px;'><div><strong>该工位已有他人预约或占用，你无法释放。</strong></div><div>只有预约人本人再次登录后，系统才允许操作。</div></div>{% endif %}</form></div></div></body></html>
"""

MY_RECORDS_TEMPLATE = """
<!DOCTYPE html><html lang='zh-CN'><head><meta charset='UTF-8'><meta name='viewport' content='width=device-width, initial-scale=1.0'><title>我的记录</title>{{ css|safe }}</head>
<body><div class='container'><div class='page-head'><div><div class='pill'>个人记录</div><div class='title' style='margin-top:8px;'>我的工位预约记录</div><div class='subtitle'>已开始的预约可释放；未开始的预约可取消。</div></div><div style='display:flex; gap:12px; flex-wrap:wrap;'><a class='btn btn-outline' style='width:auto;' href='{{ url_for("index") }}'>返回首页</a></div></div>{% if message %}<div class='toast'>{{ message }}</div>{% endif %}<div class='card section-card'><div class='section-title'>最近记录</div><div class='table-box'><table><thead><tr><th>工位</th><th>状态</th><th>开始</th><th>结束</th><th>释放/取消时间</th><th>操作</th></tr></thead><tbody>{% for row in rows %}<tr><td>{{ row.seat_display }}</td><td>{{ row.status_label }}</td><td>{{ row.start_at_display }}</td><td>{{ row.use_end_at_display }}</td><td>{{ row.released_at_display }}</td><td>{% if row.can_release %}<form method='post' action='{{ url_for("manage_my_reservation", reservation_id=row.id) }}' style='margin:0;'><input type='hidden' name='action' value='release'><button class='btn btn-danger' style='width:auto; padding:8px 12px;'>释放当前工位</button></form>{% elif row.can_cancel %}<form method='post' action='{{ url_for("manage_my_reservation", reservation_id=row.id) }}' style='margin:0;'><input type='hidden' name='action' value='cancel'><button class='btn btn-outline' style='width:auto; padding:8px 12px;'>取消预约</button></form>{% else %}-{% endif %}</td></tr>{% endfor %}{% if not rows %}<tr><td colspan='6'>暂无记录</td></tr>{% endif %}</tbody></table></div></div></div></body></html>
"""


def get_conn():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def midnight_of(date_obj: datetime) -> datetime:
    return datetime(date_obj.year, date_obj.month, date_obj.day) + timedelta(days=1)


def format_dt(dt_str: str | None) -> str:
    if not dt_str:
        return "-"
    return datetime.fromisoformat(dt_str).strftime("%Y-%m-%d %H:%M")


def default_start_str() -> str:
    return datetime.now().strftime("%Y-%m-%dT%H:%M")


def default_end_str() -> str:
    start = datetime.now()
    end = start + timedelta(hours=2)
    day_end = midnight_of(start)
    if end > day_end:
        end = day_end
    return end.strftime("%Y-%m-%dT%H:%M")


def role_from_excel(value: str) -> str:
    value = (value or "").strip()
    if value in ("管理员", "admin", "Admin", "ADMIN"):
        return "admin"
    return "employee"


def init_db():
    with LOCK:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_no TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                department TEXT,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'employee',
                status TEXT NOT NULL DEFAULT 'active',
                created_at TEXT NOT NULL
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS seats (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                seat_code TEXT UNIQUE NOT NULL,
                seat_name TEXT,
                status TEXT NOT NULL DEFAULT 'free',
                created_at TEXT NOT NULL
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS reservations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                seat_id INTEGER NOT NULL,
                user_id INTEGER NOT NULL,
                status TEXT NOT NULL,
                start_at TEXT NOT NULL,
                reserved_at TEXT NOT NULL,
                use_end_at TEXT NOT NULL,
                auto_release_at TEXT NOT NULL,
                released_at TEXT,
                release_type TEXT,
                FOREIGN KEY(seat_id) REFERENCES seats(id),
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
        """)
        conn.commit()

        for seat_code, seat_name in DEFAULT_SEATS:
            cur.execute("SELECT id FROM seats WHERE seat_code = ?", (seat_code,))
            if not cur.fetchone():
                cur.execute("INSERT INTO seats (seat_code, seat_name, status, created_at) VALUES (?, ?, 'free', ?)", (seat_code, seat_name, datetime.now().isoformat()))

        demo_users = [
            ("admin", "管理员", "IT部", "admin123", "admin"),
            ("1001", "张三", "营运部", "123456", "employee"),
            ("1002", "李四", "财务部", "123456", "employee"),
            ("1003", "王五", "IT部", "123456", "employee"),
        ]
        for employee_no, name, department, password, role in demo_users:
            cur.execute("SELECT id FROM users WHERE employee_no = ?", (employee_no,))
            if not cur.fetchone():
                cur.execute(
                    "INSERT INTO users (employee_no, name, department, password_hash, role, status, created_at) VALUES (?, ?, ?, ?, ?, 'active', ?)",
                    (employee_no, name, department, generate_password_hash(password), role, datetime.now().isoformat()),
                )
        conn.commit()
        conn.close()


def ensure_schema_upgrade():
    with LOCK:
        conn = get_conn()
        cur = conn.cursor()
        tables = {row[0] for row in cur.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
        if "users" in tables:
            cols = [row[1] for row in cur.execute("PRAGMA table_info(users)").fetchall()]
            if "role" not in cols:
                cur.execute("ALTER TABLE users ADD COLUMN role TEXT NOT NULL DEFAULT 'employee'")
                conn.commit()
            if "department" not in cols:
                cur.execute("ALTER TABLE users ADD COLUMN department TEXT")
                conn.commit()
        if "seats" in tables:
            cols = [row[1] for row in cur.execute("PRAGMA table_info(seats)").fetchall()]
            if "seat_name" not in cols:
                cur.execute("ALTER TABLE seats ADD COLUMN seat_name TEXT")
                conn.commit()
        if "reservations" in tables:
            cols = [row[1] for row in cur.execute("PRAGMA table_info(reservations)").fetchall()]
            if "start_at" not in cols:
                cur.execute("ALTER TABLE reservations ADD COLUMN start_at TEXT")
                conn.commit()
                cur.execute("UPDATE reservations SET start_at = reserved_at WHERE start_at IS NULL")
                conn.commit()
        conn.close()


def login_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login", next=request.path))
        return view_func(*args, **kwargs)
    return wrapper


def admin_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        user = get_current_user()
        if not user or user["role"] != "admin":
            return redirect(url_for("index", message="仅管理员可访问该页面"))
        return view_func(*args, **kwargs)
    return wrapper


def get_current_user():
    user_id = session.get("user_id")
    if not user_id:
        return None
    conn = get_conn()
    user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    conn.close()
    return user


def get_all_seat_rows():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM seats ORDER BY seat_code").fetchall()
    conn.close()
    return rows


def get_all_seat_codes() -> list[str]:
    return [row["seat_code"] for row in get_all_seat_rows()]


def get_seat_by_code(seat_code: str):
    conn = get_conn()
    row = conn.execute("SELECT * FROM seats WHERE seat_code = ?", (seat_code,)).fetchone()
    conn.close()
    return row


def seat_display(code: str, name: str | None) -> str:
    return name or code


def cleanup_expired():
    with LOCK:
        conn = get_conn()
        cur = conn.cursor()
        now_iso = datetime.now().isoformat()
        rows = cur.execute("SELECT id FROM reservations WHERE status = 'active' AND use_end_at <= ?", (now_iso,)).fetchall()
        for row in rows:
            cur.execute("UPDATE reservations SET status = 'released', released_at = ?, release_type = 'auto' WHERE id = ?", (datetime.now().isoformat(), row["id"]))
        conn.commit()
        conn.close()


def cleanup_loop():
    while True:
        try:
            cleanup_expired()
        except Exception as e:
            print("自动清理异常:", e)
        time.sleep(60)


def create_user(employee_no, name, department, password, role='employee'):
    employee_no = (employee_no or '').strip()
    name = (name or '').strip()
    department = (department or '').strip()
    password = (password or '').strip()
    role = role if role in ('employee', 'admin') else 'employee'
    if not employee_no or not name or not password:
        return False, '工号、姓名、密码不能为空'
    with LOCK:
        conn = get_conn()
        cur = conn.cursor()
        if cur.execute("SELECT 1 FROM users WHERE employee_no = ?", (employee_no,)).fetchone():
            conn.close()
            return False, f'工号 {employee_no} 已存在'
        cur.execute("INSERT INTO users (employee_no, name, department, password_hash, role, status, created_at) VALUES (?, ?, ?, ?, ?, 'active', ?)",
                    (employee_no, name, department, generate_password_hash(password), role, datetime.now().isoformat()))
        conn.commit()
        conn.close()
    return True, f'用户 {name}（{employee_no}）创建成功'


def update_user(employee_no, name, department, role):
    name = (name or '').strip()
    department = (department or '').strip()
    role = role if role in ('employee', 'admin') else 'employee'
    if not name:
        return False, '姓名不能为空'
    with LOCK:
        conn = get_conn()
        cur = conn.cursor()
        user = cur.execute("SELECT * FROM users WHERE employee_no = ?", (employee_no,)).fetchone()
        if not user:
            conn.close()
            return False, '用户不存在'
        if employee_no == 'admin' and role != 'admin':
            conn.close()
            return False, '默认管理员角色不可降级'
        cur.execute("UPDATE users SET name = ?, department = ?, role = ? WHERE employee_no = ?", (name, department, role, employee_no))
        conn.commit()
        conn.close()
    return True, f'用户 {employee_no} 信息已更新'


def toggle_user_status(employee_no):
    with LOCK:
        conn = get_conn()
        cur = conn.cursor()
        user = cur.execute("SELECT * FROM users WHERE employee_no = ?", (employee_no,)).fetchone()
        if not user:
            conn.close()
            return False, '用户不存在'
        if employee_no == 'admin':
            conn.close()
            return False, '默认管理员不可禁用'
        new_status = 'disabled' if user['status'] == 'active' else 'active'
        cur.execute("UPDATE users SET status = ? WHERE employee_no = ?", (new_status, employee_no))
        conn.commit()
        conn.close()
    return True, f'用户 {employee_no} 已{"禁用" if new_status == "disabled" else "启用"}'


def reset_user_password(employee_no, new_password):
    new_password = (new_password or '').strip()
    if not new_password:
        return False, '新密码不能为空'
    with LOCK:
        conn = get_conn()
        cur = conn.cursor()
        if not cur.execute("SELECT 1 FROM users WHERE employee_no = ?", (employee_no,)).fetchone():
            conn.close()
            return False, '用户不存在'
        cur.execute("UPDATE users SET password_hash = ? WHERE employee_no = ?", (generate_password_hash(new_password), employee_no))
        conn.commit()
        conn.close()
    return True, f'用户 {employee_no} 密码已重置'


def delete_user(employee_no):
    with LOCK:
        conn = get_conn()
        cur = conn.cursor()
        user = cur.execute("SELECT * FROM users WHERE employee_no = ?", (employee_no,)).fetchone()
        if not user:
            conn.close()
            return False, '用户不存在'
        if employee_no == 'admin':
            conn.close()
            return False, '默认管理员不可删除'
        if cur.execute("SELECT 1 FROM reservations WHERE user_id = ? AND status = 'active' AND use_end_at > ? LIMIT 1", (user['id'], datetime.now().isoformat())).fetchone():
            conn.close()
            return False, '该用户仍存在未结束或未来预约，不能删除'
        cur.execute("DELETE FROM users WHERE employee_no = ?", (employee_no,))
        conn.commit()
        conn.close()
    return True, f'用户 {employee_no} 已删除'


def list_users(name='', department='', role='', status=''):
    conn = get_conn()
    sql = "SELECT * FROM users WHERE 1=1"
    params = []
    if name:
        sql += " AND name LIKE ?"
        params.append(f"%{name}%")
    if department:
        sql += " AND COALESCE(department, '') LIKE ?"
        params.append(f"%{department}%")
    if role:
        sql += " AND role = ?"
        params.append(role)
    if status:
        sql += " AND status = ?"
        params.append(status)
    sql += " ORDER BY id DESC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    result = []
    for row in rows:
        result.append({
            'employee_no': row['employee_no'],
            'name': row['name'],
            'department': row['department'],
            'role': row['role'],
            'status': row['status'],
            'created_at_display': format_dt(row['created_at']),
        })
    return result


def create_seat(seat_name):
    seat_name = (seat_name or '').strip()
    if not seat_name:
        return False, '工位名称不能为空'
    with LOCK:
        conn = get_conn()
        cur = conn.cursor()
        exists = cur.execute("SELECT 1 FROM seats WHERE seat_name = ?", (seat_name,)).fetchone()
        if exists:
            conn.close()
            return False, f'工位名称 {seat_name} 已存在'
        count = cur.execute("SELECT COUNT(*) AS c FROM seats").fetchone()["c"]
        seat_code = f"SEAT-{count + 1:03d}"
        while cur.execute("SELECT 1 FROM seats WHERE seat_code = ?", (seat_code,)).fetchone():
            count += 1
            seat_code = f"SEAT-{count + 1:03d}"
        cur.execute("INSERT INTO seats (seat_code, seat_name, status, created_at) VALUES (?, ?, 'free', ?)", (seat_code, seat_name, datetime.now().isoformat()))
        conn.commit()
        conn.close()
    return True, f'工位 {seat_name} 创建成功'


def update_seat_name(seat_code, seat_name):
    seat_name = (seat_name or '').strip()
    with LOCK:
        conn = get_conn()
        cur = conn.cursor()
        if not cur.execute("SELECT 1 FROM seats WHERE seat_code = ?", (seat_code,)).fetchone():
            conn.close()
            return False, '工位不存在'
        cur.execute("UPDATE seats SET seat_name = ? WHERE seat_code = ?", (seat_name, seat_code))
        conn.commit()
        conn.close()
    return True, f'工位 {seat_code} 名称已更新'


def list_seats():
    result = []
    for row in get_all_seat_rows():
        result.append({
            'seat_code': row['seat_code'],
            'seat_name': row['seat_name'],
            'status': row['status'],
            'created_at_display': format_dt(row['created_at']),
        })
    return result


def get_all_seat_status():
    cleanup_expired()
    conn = get_conn()
    now_iso = datetime.now().isoformat()
    result = []
    for seat in get_all_seat_rows():
        seat_code = seat['seat_code']
        current_item = conn.execute("""
            SELECT r.id AS reservation_id, r.user_id, r.start_at, r.use_end_at, u.name AS user_name, u.employee_no
            FROM reservations r JOIN users u ON r.user_id = u.id JOIN seats s ON r.seat_id = s.id
            WHERE s.seat_code = ? AND r.status = 'active' AND r.start_at <= ? AND r.use_end_at > ?
            ORDER BY r.start_at ASC LIMIT 1
        """, (seat_code, now_iso, now_iso)).fetchone()
        next_item = conn.execute("""
            SELECT r.id AS reservation_id, r.user_id, r.start_at, r.use_end_at, u.name AS user_name, u.employee_no
            FROM reservations r JOIN users u ON r.user_id = u.id JOIN seats s ON r.seat_id = s.id
            WHERE s.seat_code = ? AND r.status = 'active' AND r.start_at > ?
            ORDER BY r.start_at ASC LIMIT 1
        """, (seat_code, now_iso)).fetchone()
        result.append({'seat_code': seat_code, 'seat_name': seat['seat_name'], 'current_item': current_item, 'next_item': next_item})
    conn.close()
    return result


def get_first_relevant_reservation_by_seat_code(seat_code):
    cleanup_expired()
    conn = get_conn()
    row = conn.execute("""
        SELECT r.*, u.name AS user_name, u.employee_no
        FROM reservations r JOIN users u ON r.user_id = u.id JOIN seats s ON r.seat_id = s.id
        WHERE s.seat_code = ? AND r.status = 'active' AND r.use_end_at > ?
        ORDER BY r.start_at ASC LIMIT 1
    """, (seat_code, datetime.now().isoformat())).fetchone()
    conn.close()
    return row


def serialize_reservation(row):
    if not row:
        return None
    now = datetime.now()
    start_at = datetime.fromisoformat(row['start_at'])
    end_at = datetime.fromisoformat(row['use_end_at'])
    phase = 'occupied' if start_at <= now < end_at else 'upcoming'
    return {
        'id': row['id'],
        'user_id': row['user_id'],
        'user_name': row['user_name'],
        'employee_no': row['employee_no'],
        'start_at_display': format_dt(row['start_at']),
        'use_end_at_display': format_dt(row['use_end_at']),
        'phase': phase,
    }


def build_qr_data_uri(text):
    qr = qrcode.QRCode(version=4, error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=10, border=4)
    qr.add_data(text)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    buf = BytesIO()
    img.save(buf, format='PNG')
    encoded = base64.b64encode(buf.getvalue()).decode('utf-8')
    return f'data:image/png;base64,{encoded}'


def build_seat_schedule_days(seat_code, current_user_id):
    conn = get_conn()
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    end_day = today + timedelta(days=7)
    rows = conn.execute("""
        SELECT r.*, u.name AS user_name, u.employee_no
        FROM reservations r JOIN users u ON r.user_id = u.id JOIN seats s ON r.seat_id = s.id
        WHERE s.seat_code = ? AND r.status = 'active' AND r.use_end_at > ? AND r.start_at < ?
        ORDER BY r.start_at ASC
    """, (seat_code, today.isoformat(), end_day.isoformat())).fetchall()
    conn.close()
    days = []
    now = datetime.now()
    for i in range(7):
        day_start = today + timedelta(days=i)
        day_end = day_start + timedelta(days=1)
        items = []
        for row in rows:
            start_dt = datetime.fromisoformat(row['start_at'])
            end_dt = datetime.fromisoformat(row['use_end_at'])
            seg_start = max(start_dt, day_start)
            seg_end = min(end_dt, day_end)
            if seg_end <= seg_start:
                continue
            left = ((seg_start - day_start).total_seconds() / 86400) * 100
            width = ((seg_end - seg_start).total_seconds() / 86400) * 100
            items.append({
                'left': round(left, 4),
                'width': round(width, 4),
                'is_self': row['user_id'] == current_user_id,
                'is_current': start_dt <= now < end_dt,
                'text': f"{row['user_name']} {seg_start.strftime('%H:%M')}-{seg_end.strftime('%H:%M')}",
                'title': f"{row['user_name']}（{row['employee_no']}） {format_dt(row['start_at'])} - {format_dt(row['use_end_at'])}",
                'tip': f"预约人：{row['user_name']}（{row['employee_no']}）\n开始：{format_dt(row['start_at'])}\n结束：{format_dt(row['use_end_at'])}",
            })
        now_left = round(((now - day_start).total_seconds() / 86400) * 100, 4) if i == 0 else None
        days.append({'label': day_start.strftime('%m-%d'), 'is_today': i == 0, 'now_left': now_left, 'now_label': now.strftime('%H:%M') if i == 0 else '', 'bookings': items})
    return days


def get_admin_reservation_rows(start_date='', end_date='', name=''):
    cleanup_expired()
    conn = get_conn()
    sql = """
        SELECT r.*, s.seat_code, s.seat_name, u.name AS user_name, u.employee_no, u.department, u.role AS user_role
        FROM reservations r JOIN seats s ON r.seat_id = s.id JOIN users u ON r.user_id = u.id
        WHERE 1=1
    """
    params = []
    if start_date:
        sql += " AND r.start_at >= ?"
        params.append(start_date + 'T00:00')
    if end_date:
        sql += " AND r.start_at <= ?"
        params.append(end_date + 'T23:59')
    if name:
        sql += " AND u.name LIKE ?"
        params.append(f"%{name}%")
    sql += " ORDER BY r.id DESC LIMIT 200"
    rows_db = conn.execute(sql, params).fetchall()
    conn.close()
    rows = []
    now = datetime.now()
    for row in rows_db:
        can_force_release = False
        can_force_cancel = False
        if row['status'] == 'active':
            start_dt = datetime.fromisoformat(row['start_at'])
            end_dt = datetime.fromisoformat(row['use_end_at'])
            if start_dt <= now < end_dt:
                status_label = '已开始'
                can_force_release = True
            elif start_dt > now:
                status_label = '未开始'
                can_force_cancel = True
            else:
                status_label = '已结束'
        else:
            status_label = '自动释放' if row['release_type'] == 'auto' else ('管理员取消' if row['release_type'] == 'admin_cancel' else ('管理员释放' if row['release_type'] == 'admin_release' else ('已取消' if row['release_type'] == 'cancel' else '本人释放')))
        rows.append({
            'id': row['id'],
            'seat_display': seat_display(row['seat_code'], row['seat_name']),
            'user_name': row['user_name'],
            'employee_no': row['employee_no'],
            'department': row['department'],
            'start_at_display': format_dt(row['start_at']),
            'use_end_at_display': format_dt(row['use_end_at']),
            'status_label': status_label,
            'can_force_release': can_force_release,
            'can_force_cancel': can_force_cancel,
        })
    return rows


def get_admin_report_data(start_date='', end_date='', name='', department=''):
    conn = get_conn()
    sql = """
        SELECT r.*, s.seat_code, s.seat_name, u.name AS user_name, u.employee_no, COALESCE(u.department, '未设置') AS department, u.role AS user_role
        FROM reservations r JOIN seats s ON r.seat_id = s.id JOIN users u ON r.user_id = u.id
        WHERE 1=1
    """
    params = []
    if start_date:
        sql += " AND r.start_at >= ?"
        params.append(start_date + 'T00:00')
    if end_date:
        sql += " AND r.start_at <= ?"
        params.append(end_date + 'T23:59')
    if name:
        sql += " AND u.name LIKE ?"
        params.append(f"%{name}%")
    if department:
        sql += " AND COALESCE(u.department, '未设置') = ?"
        params.append(department)
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    total_seconds = 0
    total_reservations = 0
    dept_map, user_map = {}, {}
    seat_map = {seat_display(r['seat_code'], r['seat_name']): {'count': 0, 'seconds': 0} for r in get_all_seat_rows()}

    for row in rows:
        try:
            start_dt = datetime.fromisoformat(row['start_at'])
            end_dt = datetime.fromisoformat(row['use_end_at'])
        except Exception:
            continue
        seconds = max((end_dt - start_dt).total_seconds(), 0)
        total_seconds += seconds
        total_reservations += 1
        dept = row['department']
        dept_map.setdefault(dept, {'count': 0, 'seconds': 0})
        dept_map[dept]['count'] += 1
        dept_map[dept]['seconds'] += seconds
        user_key = (row['employee_no'], row['user_name'], dept)
        user_map.setdefault(user_key, {'count': 0, 'seconds': 0})
        user_map[user_key]['count'] += 1
        user_map[user_key]['seconds'] += seconds
        sdisp = seat_display(row['seat_code'], row['seat_name'])
        seat_map.setdefault(sdisp, {'count': 0, 'seconds': 0})
        seat_map[sdisp]['count'] += 1
        seat_map[sdisp]['seconds'] += seconds

    total_hours = round(total_seconds / 3600, 1)
    total_available_hours = max(len(get_all_seat_rows()), 1) * 24 * 7
    avg_utilization = round((total_hours / total_available_hours) * 100, 1) if total_available_hours else 0

    by_department = []
    for name, data in sorted(dept_map.items(), key=lambda x: x[1]['seconds'], reverse=True):
        hours = round(data['seconds'] / 3600, 1)
        ratio = round((hours / total_hours) * 100, 1) if total_hours else 0
        by_department.append({'name': name, 'count': data['count'], 'hours': hours, 'ratio': ratio})

    by_user = []
    for (employee_no, name, dept), data in sorted(user_map.items(), key=lambda x: x[1]['seconds'], reverse=True):
        by_user.append({'employee_no': employee_no, 'name': name, 'department': dept, 'count': data['count'], 'hours': round(data['seconds'] / 3600, 1)})

    by_seat = []
    available_hours_per_seat = 24 * 7
    for sdisp, data in sorted(seat_map.items(), key=lambda x: x[1]['seconds'], reverse=True):
        hours = round(data['seconds'] / 3600, 1)
        util = round((hours / available_hours_per_seat) * 100, 1) if available_hours_per_seat else 0
        by_seat.append({'seat_display': sdisp, 'count': data['count'], 'hours': hours, 'utilization': util})

    return {'total_reservations': total_reservations, 'total_hours': total_hours, 'avg_utilization': avg_utilization}, by_department, by_user, by_seat


@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('user_id'):
        return redirect(url_for('index'))
    message = request.args.get('message', '')
    if request.method == 'POST':
        employee_no = (request.form.get('employee_no') or '').strip()
        password = request.form.get('password') or ''
        conn = get_conn()
        user = conn.execute("SELECT * FROM users WHERE employee_no = ? AND status = 'active'", (employee_no,)).fetchone()
        conn.close()
        if not user or not check_password_hash(user['password_hash'], password):
            return render_template_string(LOGIN_TEMPLATE, css=BASE_CSS, message='工号或密码不正确')
        session['user_id'] = user['id']
        return redirect(request.args.get('next') or url_for('index'))
    return render_template_string(LOGIN_TEMPLATE, css=BASE_CSS, message=message)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login', message='你已退出登录'))


@app.route('/')
@login_required
def index():
    current_user = get_current_user()
    q = (request.args.get('q') or '').strip()
    seat_rows = get_all_seat_status()
    seats = []
    people_matches = []
    occupied = 0

    for row in seat_rows:
        current_item = None
        next_item = None
        if row['current_item']:
            occupied += 1
            current_item = {
                'user_name': row['current_item']['user_name'],
                'employee_no': row['current_item']['employee_no'],
                'start_at_display': format_dt(row['current_item']['start_at']),
                'use_end_at_display': format_dt(row['current_item']['use_end_at']),
            }
        if row['next_item']:
            next_item = {
                'user_name': row['next_item']['user_name'],
                'employee_no': row['next_item']['employee_no'],
                'start_at_display': format_dt(row['next_item']['start_at']),
                'use_end_at_display': format_dt(row['next_item']['use_end_at']),
            }
        seat_item = {
            'id': row['seat_code'],
            'display_title': seat_display(row['seat_code'], row['seat_name']),
            'current_item': current_item,
            'next_item': next_item,
        }
        seats.append(seat_item)

        if q and current_item and q.lower() in current_item['user_name'].lower():
            people_matches.append({
                'user_name': current_item['user_name'],
                'seat_display': seat_item['display_title'],
                'use_end_at_display': current_item['use_end_at_display'],
            })

    filtered_seats = seats
    if q:
        filtered_seats = []
        for seat in seats:
            text_pool = [seat['display_title']]
            if seat['current_item']:
                text_pool.append(seat['current_item']['user_name'])
                text_pool.append(seat['current_item']['employee_no'])
            joined = ' | '.join([x for x in text_pool if x])
            if q.lower() in joined.lower():
                filtered_seats.append(seat)

    return render_template_string(
        INDEX_TEMPLATE,
        css=BASE_CSS,
        current_user=current_user,
        seats=filtered_seats,
        total=len(seats),
        occupied=occupied,
        free=len(seats) - occupied,
        message=request.args.get('message', ''),
        q=q,
        people_matches=people_matches,
        filtered_count=len(filtered_seats),
    )


@app.route('/admin')
@login_required
@admin_required
def admin_home():
    return render_template_string(ADMIN_HOME_TEMPLATE, css=BASE_CSS)


@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    name = (request.args.get('name') or '').strip()
    department = (request.args.get('department') or '').strip()
    role = (request.args.get('role') or '').strip()
    status = (request.args.get('status') or '').strip()
    return render_template_string(
        ADMIN_USERS_TEMPLATE,
        css=BASE_CSS,
        users=list_users(name, department, role, status),
        popup_message=request.args.get('popup', ''),
        popup_level=request.args.get('level', 'info'),
        toast_block=toast_script_block(),
        filters={'name': name, 'department': department, 'role': role, 'status': status},
    )


@app.route('/admin/users/create', methods=['POST'])
@login_required
@admin_required
def admin_create_user():
    ok, msg = create_user(request.form.get('employee_no'), request.form.get('name'), request.form.get('department'), request.form.get('password'), request.form.get('role'))
    return redirect(url_for('admin_users', popup=msg, level='success' if ok else 'error'))


@app.route('/admin/users/<employee_no>/update', methods=['POST'])
@login_required
@admin_required
def admin_update_user(employee_no):
    ok, msg = update_user(employee_no, request.form.get('name'), request.form.get('department'), request.form.get('role'))
    return redirect(url_for('admin_users', popup=msg, level='success' if ok else 'error'))


@app.route('/admin/users/<employee_no>/toggle-status', methods=['POST'])
@login_required
@admin_required
def admin_toggle_user_status(employee_no):
    ok, msg = toggle_user_status(employee_no)
    return redirect(url_for('admin_users', popup=msg, level='success' if ok else 'error'))


@app.route('/admin/users/<employee_no>/reset-password', methods=['POST'])
@login_required
@admin_required
def admin_reset_password(employee_no):
    ok, msg = reset_user_password(employee_no, request.form.get('new_password'))
    return redirect(url_for('admin_users', popup=msg, level='success' if ok else 'error'))


@app.route('/admin/users/<employee_no>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_user(employee_no):
    ok, msg = delete_user(employee_no)
    return redirect(url_for('admin_users', popup=msg, level='success' if ok else 'error'))


@app.route('/admin/users/template')
@login_required
@admin_required
def admin_user_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'users'
    ws.append(['工号', '姓名', '部门', '密码', '角色'])
    ws.append(['1008', '赵六', '营运部', '123456', '普通员工'])
    ws.append(['1009', '钱七', '财务部', '123456', '普通员工'])
    ws.append(['admin02', '管理员2', 'IT部', 'admin123', '管理员'])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='user_import_template.xlsx')


@app.route('/admin/users/import', methods=['POST'])
@login_required
@admin_required
def admin_import_users():
    file = request.files.get('file')
    if not file or not file.filename:
        return redirect(url_for('admin_users', popup='请先选择 Excel 文件', level='error'))
    try:
        wb = load_workbook(filename=BytesIO(file.read()))
        ws = wb.active
    except Exception:
        return redirect(url_for('admin_users', popup='Excel 解析失败，请检查文件格式', level='error'))
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return redirect(url_for('admin_users', popup='Excel 内容为空', level='error'))
    headers = [str(x).strip() if x is not None else '' for x in rows[0]]
    if headers[:5] != ['工号', '姓名', '部门', '密码', '角色']:
        return redirect(url_for('admin_users', popup='Excel 列头必须为 工号、姓名、部门、密码、角色', level='error'))
    success_count = 0
    fail_msgs = []
    for idx, row in enumerate(rows[1:], start=2):
        employee_no = str(row[0] or '').strip() if len(row) > 0 else ''
        name = str(row[1] or '').strip() if len(row) > 1 else ''
        department = str(row[2] or '').strip() if len(row) > 2 else ''
        password = str(row[3] or '').strip() if len(row) > 3 else ''
        role = role_from_excel(str(row[4] or '').strip()) if len(row) > 4 else 'employee'
        if not any([employee_no, name, department, password]):
            continue
        ok, msg = create_user(employee_no, name, department, password, role)
        if ok:
            success_count += 1
        else:
            fail_msgs.append(f'第{idx}行：{msg}')
    popup = f'导入完成，成功 {success_count} 条' if not fail_msgs else f'导入完成，成功 {success_count} 条；失败 {len(fail_msgs)} 条。首个失败：{fail_msgs[0]}'
    return redirect(url_for('admin_users', popup=popup, level='success' if not fail_msgs else 'info'))


@app.route('/admin/seats')
@login_required
@admin_required
def admin_seats():
    return render_template_string(ADMIN_SEATS_TEMPLATE, css=BASE_CSS, seats=list_seats(), popup_message=request.args.get('popup', ''), popup_level=request.args.get('level', 'info'), toast_block=toast_script_block())


@app.route('/admin/seats/create', methods=['POST'])
@login_required
@admin_required
def admin_create_seat():
    ok, msg = create_seat(request.form.get('seat_name'))
    return redirect(url_for('admin_seats', popup=msg, level='success' if ok else 'error'))


@app.route('/admin/seats/<seat_code>/update', methods=['POST'])
@login_required
@admin_required
def admin_update_seat(seat_code):
    ok, msg = update_seat_name(seat_code, request.form.get('seat_name'))
    return redirect(url_for('admin_seats', popup=msg, level='success' if ok else 'error'))


@app.route('/admin/reservations')
@login_required
@admin_required
def admin_reservations():
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    name = request.args.get('name', '')
    return render_template_string(ADMIN_RESERVATIONS_TEMPLATE, css=BASE_CSS, rows=get_admin_reservation_rows(start_date, end_date, name), popup_message=request.args.get('popup', ''), popup_level=request.args.get('level', 'info'), filters={'start_date': start_date, 'end_date': end_date, 'name': name}, toast_block=toast_script_block())


@app.route('/admin/reservations/<int:reservation_id>/manage', methods=['POST'])
@login_required
@admin_required
def admin_manage_reservation(reservation_id):
    action = request.form.get('action', '')
    with LOCK:
        cleanup_expired()
        conn = get_conn()
        cur = conn.cursor()
        row = cur.execute('SELECT * FROM reservations WHERE id = ?', (reservation_id,)).fetchone()
        if not row:
            conn.close()
            return redirect(url_for('admin_reservations', popup='未找到对应预约记录', level='error'))
        if row['status'] != 'active':
            conn.close()
            return redirect(url_for('admin_reservations', popup='该预约已结束或已处理', level='error'))
        now = datetime.now()
        start_dt = datetime.fromisoformat(row['start_at'])
        end_dt = datetime.fromisoformat(row['use_end_at'])
        if action == 'force_release':
            if not (start_dt <= now < end_dt):
                conn.close()
                return redirect(url_for('admin_reservations', popup='只有已开始的预约才能强制释放', level='error'))
            cur.execute("UPDATE reservations SET status = 'released', released_at = ?, release_type = 'admin_release' WHERE id = ?", (now.isoformat(), reservation_id))
            conn.commit()
            conn.close()
            return redirect(url_for('admin_reservations', popup='管理员已强制释放该预约', level='success'))
        if action == 'force_cancel':
            if not (start_dt > now):
                conn.close()
                return redirect(url_for('admin_reservations', popup='只有未开始的预约才能强制取消', level='error'))
            cur.execute("UPDATE reservations SET status = 'released', released_at = ?, release_type = 'admin_cancel' WHERE id = ?", (now.isoformat(), reservation_id))
            conn.commit()
            conn.close()
            return redirect(url_for('admin_reservations', popup='管理员已强制取消该预约', level='success'))
        conn.close()
        return redirect(url_for('admin_reservations', popup='无效操作', level='error'))


@app.route('/admin/reports')
@login_required
@admin_required
def admin_reports():
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    name = request.args.get('name', '')
    department = request.args.get('department', '')
    summary, by_department, by_user, by_seat = get_admin_report_data(start_date, end_date, name, department)
    return render_template_string(ADMIN_REPORTS_TEMPLATE, css=BASE_CSS, summary=summary, by_department=by_department, by_user=by_user, by_seat=by_seat, filters={'start_date': start_date, 'end_date': end_date, 'name': name, 'department': department})


@app.route('/my-records')
@login_required
def my_records():
    current_user = get_current_user()
    conn = get_conn()
    rows_db = conn.execute("""
        SELECT r.*, s.seat_code, s.seat_name
        FROM reservations r JOIN seats s ON r.seat_id = s.id
        WHERE r.user_id = ? ORDER BY r.id DESC LIMIT 50
    """, (current_user['id'],)).fetchall()
    conn.close()
    rows = []
    now = datetime.now()
    for row in rows_db:
        can_release = False
        can_cancel = False
        if row['status'] == 'active':
            start_dt = datetime.fromisoformat(row['start_at'])
            end_dt = datetime.fromisoformat(row['use_end_at'])
            if start_dt <= now < end_dt:
                status_label = '已开始'
                can_release = True
            elif start_dt > now:
                status_label = '未开始'
                can_cancel = True
            else:
                status_label = '已结束'
        else:
            status_label = '自动释放' if row['release_type'] == 'auto' else ('已取消' if row['release_type'] == 'cancel' else '本人释放')
        rows.append({'id': row['id'], 'seat_display': seat_display(row['seat_code'], row['seat_name']), 'status_label': status_label, 'start_at_display': format_dt(row['start_at']), 'use_end_at_display': format_dt(row['use_end_at']), 'released_at_display': format_dt(row['released_at']), 'can_release': can_release, 'can_cancel': can_cancel})
    return render_template_string(MY_RECORDS_TEMPLATE, css=BASE_CSS, rows=rows, message=request.args.get('message', ''))


@app.route('/my-records/<int:reservation_id>/manage', methods=['POST'])
@login_required
def manage_my_reservation(reservation_id):
    current_user = get_current_user()
    action = request.form.get('action', '')
    with LOCK:
        cleanup_expired()
        conn = get_conn()
        cur = conn.cursor()
        row = cur.execute('SELECT * FROM reservations WHERE id = ? AND user_id = ?', (reservation_id, current_user['id'])).fetchone()
        if not row:
            conn.close()
            return redirect(url_for('my_records', message='未找到对应预约记录'))
        if row['status'] != 'active':
            conn.close()
            return redirect(url_for('my_records', message='该预约已结束或已处理'))
        now = datetime.now()
        start_dt = datetime.fromisoformat(row['start_at'])
        end_dt = datetime.fromisoformat(row['use_end_at'])
        if action == 'release':
            if not (start_dt <= now < end_dt):
                conn.close()
                return redirect(url_for('my_records', message='只有已开始的预约才能释放'))
            cur.execute("UPDATE reservations SET status = 'released', released_at = ?, release_type = 'self' WHERE id = ?", (now.isoformat(), reservation_id))
            conn.commit()
            conn.close()
            return redirect(url_for('my_records', message='当前工位已释放'))
        if action == 'cancel':
            if not (start_dt > now):
                conn.close()
                return redirect(url_for('my_records', message='只有未开始的预约才能取消'))
            cur.execute("UPDATE reservations SET status = 'released', released_at = ?, release_type = 'cancel' WHERE id = ?", (now.isoformat(), reservation_id))
            conn.commit()
            conn.close()
            return redirect(url_for('my_records', message='预约已取消'))
        conn.close()
        return redirect(url_for('my_records', message='无效操作'))


@app.route('/qr-print')
@login_required
@admin_required
def qr_print_page():
    seats = []
    base = request.host_url.rstrip('/')
    for seat in get_all_seat_rows():
        url = f"{base}{url_for('seat_page', seat_id=seat['seat_code'])}"
        seats.append({'code': seat['seat_code'], 'name': seat['seat_name'], 'url': url, 'qr_data_uri': build_qr_data_uri(url)})
    return render_template_string(QR_PRINT_TEMPLATE, css=BASE_CSS, seats=seats)


@app.route('/seat/<seat_id>', methods=['GET', 'POST'])
@login_required
def seat_page(seat_id):
    if seat_id not in get_all_seat_codes():
        abort(404)
    current_user = get_current_user()
    seat_row = get_seat_by_code(seat_id)
    seat_display_title = seat_display(seat_id, seat_row['seat_name'] if seat_row else None)
    message = request.args.get('message', '')
    popup_message = request.args.get('popup', '')
    popup_level = request.args.get('level', 'info')
    default_start_at = default_start_str()
    default_use_end_at = default_end_str()

    def render_page(message_text='', popup_text='', popup_level_text='error', start_value=None, end_value=None):
        reservation = get_first_relevant_reservation_by_seat_code(seat_id)
        schedule_days = build_seat_schedule_days(seat_id, current_user['id'])
        return render_template_string(SEAT_TEMPLATE, css=BASE_CSS, seat_id=seat_id, seat_display_title=seat_display_title, current_user=current_user, reservation=serialize_reservation(reservation), message=message_text, popup_message=popup_text, popup_level=popup_level_text, default_start_at=start_value or default_start_at, default_use_end_at=end_value or default_use_end_at, schedule_days=schedule_days, hours=[f"{str(i).zfill(2)}:00" for i in range(24)], toast_block=toast_script_block())

    if request.method == 'POST':
        action = request.form.get('action', '')
        posted_start = request.form.get('start_at', '') or default_start_at
        posted_end = request.form.get('use_end_at', '') or default_use_end_at
        with LOCK:
            cleanup_expired()
            conn = get_conn()
            cur = conn.cursor()
            seat = cur.execute('SELECT * FROM seats WHERE seat_code = ?', (seat_id,)).fetchone()
            relevant = cur.execute("""
                SELECT r.*, u.name AS user_name, u.employee_no
                FROM reservations r JOIN users u ON r.user_id = u.id
                WHERE r.seat_id = ? AND r.status = 'active' AND r.use_end_at > ?
                ORDER BY r.start_at ASC LIMIT 1
            """, (seat['id'], datetime.now().isoformat())).fetchone()
            if action == 'reserve':
                if not posted_start or not posted_end:
                    conn.close()
                    return render_page('请选择预约开始时间和结束时间', '预约失败：请选择预约开始时间和结束时间', 'error', posted_start, posted_end)
                try:
                    start_at = datetime.fromisoformat(posted_start)
                    end_at = datetime.fromisoformat(posted_end)
                except ValueError:
                    conn.close()
                    return render_page('时间格式不正确', '预约失败：时间格式不正确', 'error', posted_start, posted_end)
                if end_at <= start_at:
                    conn.close()
                    return render_page('结束时间必须晚于开始时间', '预约失败：结束时间必须晚于开始时间', 'error', posted_start, posted_end)
                if end_at > midnight_of(start_at):
                    conn.close()
                    return render_page('结束时间不能超过预约开始当天24:00', '预约失败：结束时间不能超过预约开始当天24:00', 'error', posted_start, posted_end)
                conflict = cur.execute("""
                    SELECT r.*, u.name AS user_name, u.employee_no
                    FROM reservations r JOIN users u ON r.user_id = u.id
                    WHERE r.seat_id = ? AND r.status = 'active' AND r.use_end_at > ?
                      AND ? < r.use_end_at AND ? > r.start_at
                    ORDER BY r.start_at ASC LIMIT 1
                """, (seat['id'], datetime.now().isoformat(), start_at.isoformat(), end_at.isoformat())).fetchone()
                if conflict:
                    conn.close()
                    return render_page(f"该时间段已被 {conflict['user_name']}（{conflict['employee_no']}）预约", f"预约失败：该时间段已被 {conflict['user_name']}（{conflict['employee_no']}）预约", 'error', posted_start, posted_end)
                cur.execute("INSERT INTO reservations (seat_id, user_id, status, start_at, reserved_at, use_end_at, auto_release_at) VALUES (?, ?, 'active', ?, ?, ?, ?)", (seat['id'], current_user['id'], start_at.isoformat(), datetime.now().isoformat(), end_at.isoformat(), midnight_of(start_at).isoformat()))
                conn.commit()
                conn.close()
                return redirect(url_for('seat_page', seat_id=seat_id, message='预约成功', popup='预约成功', level='success'))
            elif action == 'release':
                if not relevant:
                    conn.close()
                    return render_page('该工位当前没有可释放的预约', '操作失败：该工位当前没有可释放的预约', 'error', posted_start, posted_end)
                now = datetime.now()
                start_dt = datetime.fromisoformat(relevant['start_at'])
                end_dt = datetime.fromisoformat(relevant['use_end_at'])
                if relevant['user_id'] != current_user['id']:
                    conn.close()
                    return render_page(f"该预约属于 {relevant['user_name']}（{relevant['employee_no']}），你不是本人，无法释放", f"操作失败：该预约属于 {relevant['user_name']}，你不是本人，无法释放", 'error', posted_start, posted_end)
                if not (start_dt <= now < end_dt):
                    conn.close()
                    return render_page('当前这条预约尚未开始，请去“我的记录”中取消预约', '操作失败：当前这条预约尚未开始，请去“我的记录”中取消预约', 'error', posted_start, posted_end)
                cur.execute("UPDATE reservations SET status = 'released', released_at = ?, release_type = 'self' WHERE id = ?", (now.isoformat(), relevant['id']))
                conn.commit()
                conn.close()
                return redirect(url_for('seat_page', seat_id=seat_id, message='本人释放成功', popup='本人释放成功', level='success'))
            conn.close()
    return render_page(message, popup_message, popup_level, default_start_at, default_use_end_at)


if __name__ == '__main__':
    init_db()
    ensure_schema_upgrade()
    t = threading.Thread(target=cleanup_loop, daemon=True)
    t.start()
    app.run(host='0.0.0.0', port=5000, debug=True)
